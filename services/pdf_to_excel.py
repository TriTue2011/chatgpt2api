"""Chuyển PDF → Excel (.xlsx) — pipeline nhiều tầng (tham khảo open-source).

Tham khảo kỹ thuật (không vendor lock 1 repo):
  - noworneverev/PDF2Excel ………… pdfplumber.extract_tables → sheet
  - monambike/pdfconverter-… …… làm sạch bảng: drop dòng/cột rỗng, gộp xuống dòng trong ô
  - TheTechTiger/MHTCET-… ……… fingerprint bảng trùng (header lặp), bỏ sheet trùng
  - shine-jayakumar/Extract-… … pdftotext -layout + tách cột khoảng trắng

Thứ tự ưu tiên (mỗi tầng chỉ chạy nếu tầng trước không ra bảng “đủ tốt”):
  1. pdfplumber tables (nếu cài)
  2. PyMuPDF page.find_tables()
  3. Text layout: pdftotext -layout / fitz words → cột theo khoảng trắng
  4. OCR markdown tables (scan) qua pdf_intent.extract_markdown

Xuất .xlsx: openpyxl nếu có, không thì OOXML zip tối thiểu.
"""
from __future__ import annotations

import hashlib
import logging
import re
import subprocess
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

logger = logging.getLogger(__name__)

# Bảng tối thiểu bao nhiêu ô có chữ mới coi là “có dữ liệu”
_MIN_NONEMPTY_CELLS = 4
_MAX_PAGES = 80
_MAX_ROWS = 8000
_MAX_COLS = 40


def convert_pdf_to_xlsx(pdf_path: str, out_path: str | None = None) -> dict[str, Any]:
    """PDF → .xlsx. Trả {ok, path, method, sheets, error, strategies}."""
    src = Path(pdf_path)
    if not src.is_file():
        return {"ok": False, "error": f"không thấy PDF: {pdf_path}"}
    dest = Path(out_path) if out_path else src.with_suffix(".xlsx")

    tried: list[str] = []
    sheets: list[tuple[str, list[list[str]]]] = []
    method = ""

    # ── 1) pdfplumber (PDF2Excel) ─────────────────────────────────────────
    sheets, method = _extract_pdfplumber(src)
    if sheets:
        tried.append(f"pdfplumber:{len(sheets)}")
    else:
        tried.append("pdfplumber:0")

    # ── 2) PyMuPDF find_tables ────────────────────────────────────────────
    if not _tables_good(sheets):
        s2, m2 = _extract_pymupdf_tables(src)
        tried.append(f"pymupdf_tables:{len(s2)}")
        if _tables_good(s2) or (s2 and not _tables_good(sheets)):
            sheets, method = s2, m2

    # ── 3) Layout text → cột (pdftotext -layout / fitz words) ─────────────
    if not _tables_good(sheets):
        s3, m3 = _extract_layout_columns(src)
        tried.append(f"layout:{len(s3)}")
        if s3 and (not sheets or _score(s3) > _score(sheets)):
            sheets, method = s3, m3

    # ── 4) OCR markdown tables (scan) ─────────────────────────────────────
    if not _tables_good(sheets):
        s4, m4 = _extract_ocr_markdown(src)
        tried.append(f"ocr:{len(s4)}")
        if s4 and (not sheets or _tables_good(s4)):
            sheets, method = s4, m4

    # ── 5) Last resort: plain text lines ──────────────────────────────────
    if not sheets:
        s5, m5 = _extract_plain_lines(src)
        tried.append(f"plain:{len(s5)}")
        sheets, method = s5, m5

    if not sheets:
        return {
            "ok": False,
            "error": "Không trích được bảng/chữ từ PDF",
            "strategies": tried,
        }

    # Làm sạch + bỏ bảng trùng 100% (MHTCET fingerprint)
    cleaned = _clean_all_sheets(sheets)
    if not cleaned:
        return {"ok": False, "error": "Bảng rỗng sau khi làm sạch", "strategies": tried}

    # Gộp mọi trang vào 1 sheet — báo giá/PDF nhiều trang cùng header:
    # user mở Excel thường chỉ thấy sheet đầu, tưởng mất trang 2+.
    merged = _merge_sheets_one_table(cleaned)
    to_write = merged if merged else cleaned

    try:
        _write_xlsx(dest, to_write)
    except Exception as exc:
        return {"ok": False, "error": f"ghi xlsx lỗi: {exc}", "strategies": tried}

    return {
        "ok": True,
        "path": str(dest),
        "method": method or "mixed",
        "sheets": len(to_write),
        "pages_extracted": len(cleaned),
        "strategies": tried,
    }


# ── Extractors ───────────────────────────────────────────────────────────────


def _extract_pdfplumber(src: Path) -> tuple[list[tuple[str, list[list[str]]]], str]:
    """pdfplumber.extract_tables — PDF2Excel style."""
    try:
        import pdfplumber  # type: ignore
    except Exception:
        return [], ""
    sheets: list[tuple[str, list[list[str]]]] = []
    try:
        with pdfplumber.open(str(src)) as pdf:
            for i, page in enumerate(pdf.pages[:_MAX_PAGES]):
                try:
                    tables = page.extract_tables() or []
                except Exception:
                    tables = []
                for ti, table in enumerate(tables):
                    rows = [[_cell(c) for c in (row or [])] for row in (table or [])]
                    rows = _normalize_rows(rows)
                    if _table_nonempty(rows):
                        sheets.append((f"P{i + 1}_B{ti + 1}", rows))
    except Exception as exc:
        logger.warning("pdf_to_excel pdfplumber: %s", exc)
        return [], ""
    return sheets, ("pdfplumber" if sheets else "")


def _extract_pymupdf_tables(src: Path) -> tuple[list[tuple[str, list[list[str]]]], str]:
    try:
        import fitz  # type: ignore
    except Exception:
        return [], ""
    sheets: list[tuple[str, list[list[str]]]] = []
    try:
        doc = fitz.open(src)
        try:
            for i, page in enumerate(doc):
                if i >= _MAX_PAGES:
                    break
                finder = getattr(page, "find_tables", None)
                if not callable(finder):
                    continue
                try:
                    tabs = finder()
                    tables = getattr(tabs, "tables", None) or list(tabs or [])
                except Exception:
                    tables = []
                for ti, tab in enumerate(tables):
                    try:
                        data = tab.extract() if hasattr(tab, "extract") else None
                    except Exception:
                        data = None
                    if not data:
                        continue
                    rows = [[_cell(c) for c in row] for row in data]
                    rows = _normalize_rows(rows)
                    if _table_nonempty(rows):
                        sheets.append((f"P{i + 1}_B{ti + 1}", rows))
        finally:
            doc.close()
    except Exception as exc:
        logger.warning("pdf_to_excel pymupdf tables: %s", exc)
        return [], ""
    return sheets, ("pymupdf_tables" if sheets else "")


def _extract_layout_columns(src: Path) -> tuple[list[tuple[str, list[list[str]]]], str]:
    """pdftotext -layout hoặc fitz text blocks → tách cột khoảng trắng (shine-jayakumar spirit)."""
    text = _pdftotext_layout(src)
    method = "pdftotext_layout"
    if not text:
        text = _fitz_text(src)
        method = "fitz_layout"
    if not text.strip():
        return [], ""

    sheets: list[tuple[str, list[list[str]]]] = []
    # Split by form feed if multi-page
    pages = re.split(r"\f", text)
    for i, page_txt in enumerate(pages[:_MAX_PAGES]):
        rows = _lines_to_columns(page_txt)
        if _table_nonempty(rows) and _looks_tabular(rows):
            sheets.append((f"Trang_{i + 1}", rows))
        elif rows:
            # single-col still useful
            sheets.append((f"Trang_{i + 1}", rows))
    if not sheets and text.strip():
        rows = _lines_to_columns(text)
        if rows:
            sheets.append(("Noi_dung", rows))
    return sheets, (method if sheets else "")


def _extract_ocr_markdown(src: Path) -> tuple[list[tuple[str, list[list[str]]]], str]:
    try:
        from services.pdf_intent import extract_markdown
        md = extract_markdown(str(src)) or ""
    except Exception as exc:
        logger.warning("pdf_to_excel ocr: %s", exc)
        return [], ""
    md_sheets = _sheets_from_markdown(md)
    if md_sheets:
        return md_sheets, "ocr_markdown"
    if md.strip():
        rows = [[ln] for ln in md.splitlines() if ln.strip()][:_MAX_ROWS]
        return [("Noi_dung", rows)], "ocr_text"
    return [], ""


def _extract_plain_lines(src: Path) -> tuple[list[tuple[str, list[list[str]]]], str]:
    text = _fitz_text(src) or _pdftotext_layout(src)
    if not text.strip():
        return [], ""
    rows = [[ln] for ln in text.splitlines() if ln.strip()][:_MAX_ROWS]
    return ([("Noi_dung", rows)], "plain_text") if rows else ([], "")


# ── Text helpers ─────────────────────────────────────────────────────────────


def _pdftotext_layout(src: Path) -> str:
    try:
        r = subprocess.run(
            ["pdftotext", "-layout", "-enc", "UTF-8", str(src), "-"],
            capture_output=True, text=True, timeout=60,
        )
        if r.returncode == 0 and (r.stdout or "").strip():
            return r.stdout
    except Exception as exc:
        logger.debug("pdftotext: %s", exc)
    return ""


def _fitz_text(src: Path) -> str:
    try:
        import fitz  # type: ignore
        doc = fitz.open(src)
        try:
            parts = []
            for i, page in enumerate(doc):
                if i >= _MAX_PAGES:
                    break
                parts.append(page.get_text("text") or "")
            return "\f".join(parts)
        finally:
            doc.close()
    except Exception:
        return ""


def _lines_to_columns(page_txt: str) -> list[list[str]]:
    """Tách dòng thành cột theo 2+ khoảng trắng (layout text tables)."""
    rows: list[list[str]] = []
    for ln in (page_txt or "").splitlines():
        s = ln.rstrip()
        if not s.strip():
            continue
        # 2+ spaces as delimiter (common fixed-width PDF tables)
        if re.search(r"\s{2,}", s.strip()):
            cells = [c.strip() for c in re.split(r"\s{2,}", s.strip()) if c.strip()]
        else:
            # tab
            if "\t" in s:
                cells = [c.strip() for c in s.split("\t") if c.strip()]
            else:
                cells = [s.strip()]
        if cells:
            rows.append(cells[:_MAX_COLS])
        if len(rows) >= _MAX_ROWS:
            break
    return _normalize_rows(rows)


def _looks_tabular(rows: list[list[str]]) -> bool:
    if len(rows) < 2:
        return False
    multi = sum(1 for r in rows if len(r) >= 2)
    return multi >= max(2, len(rows) // 3)


# ── Cleaning (monambike + MHTCET) ────────────────────────────────────────────


def _clean_all_sheets(
    sheets: list[tuple[str, list[list[str]]]],
) -> list[tuple[str, list[list[str]]]]:
    out: list[tuple[str, list[list[str]]]] = []
    seen_fp: set[str] = set()
    for name, rows in sheets:
        rows = _clean_table(rows)
        if not _table_nonempty(rows):
            continue
        fp = _table_fingerprint(rows)
        if fp in seen_fp:
            # header/table lặp y hệt giữa các trang
            continue
        seen_fp.add(fp)
        out.append((name, rows))
    return out


def _header_signature(row: list[str]) -> str:
    """Chuẩn hóa header để so khớp lặp giữa các trang (STT/Danh mục/…)."""
    parts = []
    for c in row or []:
        t = re.sub(r"\s+", " ", (c or "").strip().lower())
        parts.append(t)
    return "|".join(parts)


def _merge_sheets_one_table(
    sheets: list[tuple[str, list[list[str]]]],
) -> list[tuple[str, list[list[str]]]]:
    """Gộp nhiều sheet trang (cùng kiểu bảng) thành 1 sheet liên tục.

    - Giữ header của sheet đầu.
    - Sheet sau: bỏ dòng header lặp (cùng signature).
    - Pad cột cho đồng rộng.
    """
    if len(sheets) <= 1:
        return list(sheets)

    # Chỉ gộp nếu ≥2 sheet có ≥2 cột (bảng), không gộp plain text 1 cột
    tabular = [(n, r) for n, r in sheets if r and max(len(x) for x in r) >= 2]
    if len(tabular) <= 1:
        return list(sheets)

    header = list(tabular[0][1][0]) if tabular[0][1] else []
    hdr_sig = _header_signature(header)
    width = max(len(r) for _, rows in tabular for r in rows) if tabular else len(header)
    if header:
        header = header + [""] * (width - len(header))

    body: list[list[str]] = []
    for _name, rows in tabular:
        if not rows:
            continue
        start = 0
        if hdr_sig and _header_signature(rows[0]) == hdr_sig:
            start = 1
        for row in rows[start:]:
            padded = list(row) + [""] * (width - len(row))
            body.append(padded[:width])

    if not body and not header:
        return list(sheets)

    out_rows: list[list[str]] = []
    if header:
        out_rows.append(header[:width])
    out_rows.extend(body)
    # Tên sheet rõ: Bang_day_du (all pages)
    return [("Bang_day_du", out_rows[:_MAX_ROWS])]


def _clean_table(rows: list[list[str]]) -> list[list[str]]:
    """Drop empty rows/cols; giữ xuống dòng trong ô (mô tả kỹ thuật PDF)."""
    if not rows:
        return []
    # Giữ newline trong ô (Excel wrap text); chỉ gọn space ngang + strip mép.
    cleaned = []
    for row in rows:
        cells = []
        for c in row:
            s = _cell(c)
            # Chuẩn hóa CRLF; bỏ khoảng trắng thừa hai bên mỗi dòng
            lines = [ln.strip() for ln in s.replace("\r\n", "\n").replace("\r", "\n").split("\n")]
            # Bỏ dòng trống đầu/cuối nhưng giữ bullet giữa chừng
            while lines and not lines[0]:
                lines.pop(0)
            while lines and not lines[-1]:
                lines.pop()
            cells.append("\n".join(lines).strip())
        cleaned.append(cells)
    # drop fully empty rows
    cleaned = [r for r in cleaned if any(c.strip() for c in r)]
    if not cleaned:
        return []
    # pad columns
    width = max(len(r) for r in cleaned)
    cleaned = [r + [""] * (width - len(r)) for r in cleaned]
    # drop empty columns
    keep_cols = [
        j for j in range(width)
        if any((cleaned[i][j] or "").strip() for i in range(len(cleaned)))
    ]
    if not keep_cols:
        return []
    cleaned = [[r[j] for j in keep_cols] for r in cleaned]
    # drop repeated header-only rows (same as first row)
    if len(cleaned) > 2:
        header = cleaned[0]
        body = [r for r in cleaned[1:] if r != header]
        cleaned = [header] + body
    return cleaned[:_MAX_ROWS]


def _table_fingerprint(rows: list[list[str]]) -> str:
    """Fingerprint để bỏ bảng trùng (MHTCET)."""
    sample = rows[:8]
    blob = "||".join("|".join(c for c in r) for r in sample)
    return hashlib.sha1(blob.encode("utf-8", errors="ignore"), usedforsecurity=False).hexdigest()[:16]


def _normalize_rows(rows: list[list[str]]) -> list[list[str]]:
    out = []
    for row in rows:
        if row is None:
            continue
        out.append([_cell(c) for c in row][:_MAX_COLS])
    return out[:_MAX_ROWS]


def _table_nonempty(rows: list[list[str]]) -> bool:
    n = sum(1 for r in rows for c in r if (c or "").strip())
    return n >= _MIN_NONEMPTY_CELLS


def _tables_good(sheets: list[tuple[str, list[list[str]]]]) -> bool:
    if not sheets:
        return False
    # Ít nhất 1 sheet có ≥2 cột hoặc ≥ _MIN_NONEMPTY_CELLS ô
    for _, rows in sheets:
        if any(len(r) >= 2 for r in rows) and _table_nonempty(rows):
            return True
        if _table_nonempty(rows) and len(rows) >= 3:
            return True
    return False


def _score(sheets: list[tuple[str, list[list[str]]]]) -> int:
    score = 0
    for _, rows in sheets:
        cols = max((len(r) for r in rows), default=0)
        cells = sum(1 for r in rows for c in r if (c or "").strip())
        score += cells + cols * 10 + len(rows)
    return score


def _cell(v: object) -> str:
    if v is None:
        return ""
    return str(v).replace("\x00", "").strip()


def _sheets_from_markdown(md: str) -> list[tuple[str, list[list[str]]]]:
    lines = (md or "").splitlines()
    tables: list[list[list[str]]] = []
    cur: list[list[str]] = []
    for ln in lines:
        s = ln.strip()
        if s.startswith("|") and s.endswith("|"):
            if re.match(r"^\|[\s\-:|]+\|$", s):
                continue
            cells = [c.strip() for c in s.strip("|").split("|")]
            cur.append(cells)
        else:
            if cur:
                tables.append(cur)
                cur = []
    if cur:
        tables.append(cur)
    out: list[tuple[str, list[list[str]]]] = []
    for i, t in enumerate(tables):
        t = _normalize_rows(t)
        if _table_nonempty(t):
            out.append((f"Bang_{i + 1}", t))
    return out


# ── XLSX writers ─────────────────────────────────────────────────────────────


def _write_xlsx(path: Path, sheets: list[tuple[str, list[list[str]]]]) -> None:
    try:
        from openpyxl import Workbook  # type: ignore

        from openpyxl.styles import Alignment, Font  # type: ignore

        wb = Workbook()
        default = wb.active
        first = True
        wrap = Alignment(wrap_text=True, vertical="top")
        # Font chuẩn VN cho mọi cell (Times New Roman 13) — theo yêu cầu user.
        font13 = Font(name="Times New Roman", size=13)
        for name, rows in sheets:
            safe = _safe_sheet_name(name)
            if first:
                ws = default
                ws.title = safe
                first = False
            else:
                ws = wb.create_sheet(safe)
            for r_i, row in enumerate(rows, start=1):
                for c_i, val in enumerate(row, start=1):
                    coerced = _coerce_value(val)
                    cell = ws.cell(r_i, c_i, coerced)
                    cell.font = font13
                    if isinstance(coerced, str) and "\n" in coerced:
                        cell.alignment = wrap
            try:
                ws.column_dimensions["B"].width = 56
            except Exception:
                pass
        if first:
            default["A1"] = "(trống)"
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return
    except Exception as exc:
        logger.info("openpyxl unavailable (%s) — minimal xlsx", exc)
    _write_xlsx_minimal(path, sheets)


def _coerce_value(val: str) -> object:
    """Giữ số tiền VN (1.234.567 / 938.000) dạng text — không ép float.

    Trước đây ``938.000`` → float 938.0 mất 3 số 0 (dấu chấm = hàng nghìn VN).
    Chỉ ép int thuần (không dấu chấm/phẩy) hoặc decimal rõ (1 dấu + phần lẻ ≠ 3 số
    kiểu nghìn VN lặp).
    """
    s = (val or "").strip()
    if not s:
        return ""
    # Giữ nguyên nếu có xuống dòng (mô tả kỹ thuật)
    if "\n" in s:
        return s
    # Số kiểu Việt Nam: 1.234 hoặc 1.234.567 hoặc 938.000 → giữ text
    if re.fullmatch(r"-?\d{1,3}(\.\d{3})+", s):
        return s
    # Biến thể dùng dấu phẩy hàng nghìn: 1,234,567
    if re.fullmatch(r"-?\d{1,3}(,\d{3})+", s):
        return s
    # int thuần
    if re.fullmatch(r"-?\d{1,15}", s):
        try:
            return int(s)
        except Exception:
            return s
    # Decimal thực (một dấu): 12.5 / 12,5 — KHÔNG phải xxx.000 hàng nghìn
    if re.fullmatch(r"-?\d{1,12}[.,]\d{1,8}", s):
        frac = re.split(r"[.,]", s)[-1]
        # 3 chữ số sau dấu + phần nguyên ≤3 số → hay là hàng nghìn VN (938.000)
        if len(frac) == 3 and re.fullmatch(r"-?\d{1,3}", re.split(r"[.,]", s)[0]):
            return s
        try:
            return float(s.replace(",", "."))
        except Exception:
            return s
    return s


def _safe_sheet_name(name: str) -> str:
    s = re.sub(r"[\[\]\*\/\\?:]", "_", (name or "Sheet")[:31]) or "Sheet"
    return s


def _write_xlsx_minimal(path: Path, sheets: list[tuple[str, list[list[str]]]]) -> None:
    used: set[str] = set()
    named: list[tuple[str, list[list[str]]]] = []
    for name, rows in sheets:
        base = _safe_sheet_name(name)
        n = base
        i = 1
        while n.lower() in used:
            i += 1
            n = _safe_sheet_name(f"{base[:28]}_{i}")
        used.add(n.lower())
        named.append((n, rows))

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", _content_types(len(named)))
        z.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>""",
        )
        z.writestr("xl/workbook.xml", _workbook_xml(named))
        z.writestr("xl/_rels/workbook.xml.rels", _workbook_rels(len(named)))
        z.writestr(
            "xl/styles.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills count="1"><fill><patternFill patternType="none"/></fill></fills>
  <borders count="1"><border/></borders>
  <cellStyleXfs count="1"><xf/></cellStyleXfs>
  <cellXfs count="1"><xf/></cellXfs>
</styleSheet>""",
        )
        for i, (_, rows) in enumerate(named, start=1):
            z.writestr(f"xl/worksheets/sheet{i}.xml", _sheet_xml(rows))
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(buf.getvalue())


def _content_types(n: int) -> str:
    overrides = [
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>',
    ]
    for i in range(1, n + 1):
        overrides.append(
            f'<Override PartName="/xl/worksheets/sheet{i}.xml" '
            f'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        + "".join(overrides)
        + "</Types>"
    )


def _workbook_xml(named: list[tuple[str, list[list[str]]]]) -> str:
    sheets = []
    for i, (name, _) in enumerate(named, start=1):
        sheets.append(f'<sheet name="{escape(name)}" sheetId="{i}" r:id="rId{i}"/>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets>{"".join(sheets)}</sheets></workbook>'
    )


def _workbook_rels(n: int) -> str:
    rels = []
    for i in range(1, n + 1):
        rels.append(
            f'<Relationship Id="rId{i}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{i}.xml"/>'
        )
    rels.append(
        f'<Relationship Id="rId{n + 1}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
        'Target="styles.xml"/>'
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + "".join(rels)
        + "</Relationships>"
    )


def _col_name(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s or "A"


def _sheet_xml(rows: list[list[str]]) -> str:
    parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>',
    ]
    for r_i, row in enumerate(rows[:_MAX_ROWS], start=1):
        cells = []
        for c_i, val in enumerate(row[:_MAX_COLS], start=1):
            ref = f"{_col_name(c_i)}{r_i}"
            t = escape(val or "")
            cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{t}</t></is></c>')
        parts.append(f'<row r="{r_i}">{"".join(cells)}</row>')
    parts.append("</sheetData></worksheet>")
    return "".join(parts)
